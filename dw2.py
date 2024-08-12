import os
import pandas as pd
import requests
from urllib.parse import urlparse  # 导入urlparse
from openpyxl import load_workbook

# 获取当前脚本所在目录
current_dir = os.path.dirname(os.path.abspath(__file__))

# 查找目录中以 ".xlsx" 结尾的文件
excel_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]

# 如果目录中存在Excel文件，则选择第一个文件
if len(excel_files) > 0:
    excel_file = os.path.join(current_dir, excel_files[0])
    print(f"读取文件: {excel_file}")

    # 使用openpyxl加载工作簿
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # 获取DataFrame用于处理数据（除了超链接外的内容）
    df = pd.read_excel(excel_file)

    # 从第二行开始，遍历Excel表格
    for index, row in df.iterrows():
        if index == 0:  # 跳过标题行
            continue

        # 读取第11列的值并以此值建立文件夹
        folder_name = str(row[10])  # 第11列是索引10
        folder_path = os.path.join(current_dir, folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # 获取第1列中的超链接（不是显示的文本）
        cell = sheet.cell(row=index+2, column=1)  # +2因为Openpyxl的行索引从1开始，跳过标题行
        hyperlink = cell.hyperlink.target if cell.hyperlink else None

        if hyperlink:
            try:
                # 下载文件
                response = requests.get(hyperlink)
                if response.status_code == 200:
                    # 获取文件名
                    url_path = urlparse(hyperlink).path
                    zip_filename = os.path.basename(url_path)

                    # 保存文件到对应的文件夹
                    save_path = os.path.join(folder_path, zip_filename)
                    with open(save_path, 'wb') as file:
                        file.write(response.content)

                    # 重命名文件为 "第1列_第3列"
                    new_filename = f"{str(row[0])} {str(row[2])}.zip"  # 第1列和第3列的值
                    new_path = os.path.join(folder_path, new_filename)
                    os.rename(save_path, new_path)

                    print(f"文件下载并重命名为: {new_path}")
                else:
                    print(f"无法下载文件: {hyperlink}, 状态码: {response.status_code}")
            except Exception as e:
                print(f"处理链接时出错: {hyperlink}, 错误: {str(e)}")
        else:
            print(f"超链接无效或不存在: {hyperlink}")
else:
    print("当前目录中没有找到Excel文件。")
