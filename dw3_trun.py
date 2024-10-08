import os
import pandas as pd
import requests
from urllib.parse import urlparse
from openpyxl import load_workbook
import re  # 导入正则表达式模块

# 获取当前脚本所在目录
current_dir = os.path.dirname(os.path.abspath(__file__))

# 查找目录中以 ".xlsx" 结尾的文件
excel_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]

# 定义最大路径长度
WINDOWS_MAX_PATH_LENGTH = 255

# 定义一个函数来清理文件名
def clean_filename(filename):
    # 替换掉不允许的字符为下划线
    filename = re.sub(r'[\\/*?:"<>|]', '_', filename)
    # 如果文件名超过Windows允许的最大长度，进行截断
    if len(filename) > WINDOWS_MAX_PATH_LENGTH:
        filename = filename[:WINDOWS_MAX_PATH_LENGTH]
    return filename

# 如果目录中存在Excel文件，则选择第一个文件
if len(excel_files) > 0:
    excel_file = os.path.join(current_dir, excel_files[0])
    print(f"读取文件: {excel_file}")

    # 使用openpyxl加载工作簿
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # 获取DataFrame用于处理数据（除了超链接外的内容）
    df = pd.read_excel(excel_file)

    # 从第二行开始，遍历Excel表格
    for index, row in df.iterrows():
        if index == 0:  # 跳过标题行
            continue

        # 读取第11列的值并以此值建立文件夹
        folder_name = clean_filename(str(row.iloc[10]))  # 使用 iloc 访问第11列
        folder_path = os.path.join(current_dir, folder_name)
        
        # 检查路径是否太长
        if len(folder_path) > WINDOWS_MAX_PATH_LENGTH:
            print(f"警告: 文件路径太长，已截断: {folder_path}")
            folder_path = folder_path[:WINDOWS_MAX_PATH_LENGTH]

        # 创建文件夹
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # 获取第1列中的超链接（不是显示的文本）
        cell = sheet.cell(row=index+2, column=1)  # +2因为Openpyxl的行索引从1开始，跳过标题行
        hyperlink = cell.hyperlink.target if cell.hyperlink else None

        if hyperlink:
            try:
                # 下载文件
                response = requests.get(hyperlink)
                if response.status_code == 200:
                    # 获取文件名
                    url_path = urlparse(hyperlink).path
                    zip_filename = os.path.basename(url_path)

                    # 保存文件到对应的文件夹
                    save_path = os.path.join(folder_path, zip_filename)
                    with open(save_path, 'wb') as file:
                        file.write(response.content)

                    # 生成新的文件名并清理非法字符
                    new_filename = clean_filename(f"{str(row.iloc[0])} {str(row.iloc[2])} {str(row.iloc[1])}.zip")
                    new_path = os.path.join(folder_path, new_filename)
                    os.rename(save_path, new_path)

                    print(f"文件下载并重命名为: {new_path}")
                else:
                    print(f"无法下载文件: {hyperlink}, 状态码: {response.status_code}")
            except Exception as e:
                print(f"处理链接时出错: {hyperlink}, 错误: {str(e)}")
        else:
            print(f"超链接无效或不存在: {hyperlink}")
else:
    print("当前目录中没有找到Excel文件。")
